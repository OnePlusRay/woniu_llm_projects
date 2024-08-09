import openpyxl
from src.table import Table
from src.merge_table import merge_base_table_dict, merge_table_n_write
from src.upload_data import upload_origin_file, upload_result_file
from src.map_factor_name import batch_map_factor_name
import os
import pandas as pd
import pdb
import time 
import requests
from src.save_table import save_base_table_dict, load_base_table_dict
from utils.checkpoint import save_checkpoint, load_checkpoint
import datetime


def get_res(file_path, output_dir_path, is_test=False):
    '''
        file_path: 待处理文件路径
        output_dir_path: 输出文件路径
        is_test: 是否为小数据集测试，若是，无需保存中间结果
    '''
    start = 0
    base_table_dict = {}

    # debug模式：尝试加载检查点
    # if is_debug:
    #     checkpoint_dir = ''
    #     data_file_path = os.path.join(checkpoint_dir, 'data.pkl')
    #     checkpoint_file_path = os.path.join(checkpoint_dir, 'checkpoint.pkl')
    #     checkpoint, base_table_dict = load_checkpoint(data_file_path, checkpoint_file_path)
    #     if checkpoint:
    #         start = checkpoint['current']  # 从上一次保存的检查点位置开始遍历表格
    #         print(f"Resuming from checkpoint: {start}")

    # 获取当前时间戳
    timestamp = int(time.time())

    # 定义保存中间检查点的文件夹和文件名
    # checkpoint_base_dir = '/sdata/chenrui/ai-project/insurance_rate/data/checkpoints'
    # checkpoint_dir = os.path.join(checkpoint_base_dir, f'checkpoint_{timestamp}')
    # checkpoint_file = os.path.join(checkpoint_dir, 'checkpoint.pkl')
    # data_file = os.path.join(checkpoint_dir, 'data.pkl')
    # os.makedirs(checkpoint_dir, exist_ok=True)  # 创建新的文件夹路径

    #获取原始excel文件url
    origin_url = upload_origin_file(file_path)

    #创建输出文件路径（加时间戳）
    workbook = openpyxl.load_workbook(file_path)
    filename = os.path.basename(file_path).replace('.xlsx',f'_flatten_{timestamp}.xlsx').replace('.csv',f'_flatten_{timestamp}.xlsx')
    output_path = os.path.join(output_dir_path, filename)
    if not os.path.exists(output_path):
        # 创建输出文件路径的目录
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
    writer = pd.ExcelWriter(output_path)
    # pdb.set_trace()


    # 按sheet处理表格，中间结果保存到base_table_dict
    for i in range(start, len(workbook.sheetnames)):
        sheet_name = workbook.sheetnames[i]

        # 获取base_name，同一个责任需要合并表格，例如：基本责任_1，基本责任_2
        if '_' in sheet_name:
            base_name = sheet_name.replace(sheet_name.split('_')[-1],'').replace('_','')
        else:
            base_name = sheet_name
        print(base_name)
        print(sheet_name)
        #pdb.set_trace()

        # 创建Table类对象
        table = Table(sheet_name)

        # 读取sheet内容
        table.load_table(workbook[sheet_name])
        # pdb.set_trace()

        # 识别表格结构
        table.recognize_table_structure()
        # pdb.set_trace()

        # 平铺表格
        table.flatten()
        
        # 合并base_table 
        base_table_dict = merge_base_table_dict(base_table_dict, base_name, table)  # base_table_dict储存了表格的全部信息

        # 每隔 5 张 sheet 保存一次检查点和数据（也可以每张都保存）
        # if (i + 1) % checkpoint_interval == 0:
        # save_checkpoint(base_table_dict, {'current': i}, data_file, checkpoint_file)

        # 保存目前的base_table_dict
        if not is_test:
            save_base_table_dict(base_table_dict, base_name)

    # 加载保存的base_table_dict
    # base_table_dict = load_base_table_dict('/sdata/chenrui/ai-project/insurance_rate/data/save_files/复星大黄蜂10号')

    # pdb.set_trace()
    # 因子名称规范化
    batch_map_factor_name(base_table_dict)
    # pdb.set_trace()

    #合并所有表格并write文件
    # pdb.set_trace()
    writer = merge_table_n_write(base_table_dict, writer)
    writer._save()  
    writer.close()
    result_url = upload_result_file(output_path)

    # 任务完成后删除检查点文件
    # if os.path.exists(checkpoint_file):
        
    #     os.remove(checkpoint_file)
    #     os.remove(data_file)
    # print("Task completed, checkpoint is deleted")
    return origin_url, result_url